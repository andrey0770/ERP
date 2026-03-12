"""Generate Excel table with cue codes for editing."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Кии — коды"

# Styles
header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="3F51B5")
cat_fill = PatternFill("solid", fgColor="E8EAF6")
cat_font = Font(bold=True, size=11, color="3F51B5")
code_font = Font(bold=True, size=12, name="Consolas", color="00838F")
sku_font = Font(bold=True, size=11, name="Consolas")
thin_border = Border(
    left=Side(style='thin', color='DDDDDD'),
    right=Side(style='thin', color='DDDDDD'),
    top=Side(style='thin', color='DDDDDD'),
    bottom=Side(style='thin', color='DDDDDD'),
)

# Headers
headers = ["Артикул", "Код товара", "Название", "Поставщик", "Конструкция", "Игра", "Старый ID"]
widths = [12, 14, 42, 14, 22, 14, 12]
for col, (h, w) in enumerate(zip(headers, widths), 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    ws.column_dimensions[chr(64+col)].width = w

# Data — only our suppliers (Мейлин, Smart, Конди)
cues = [
    # (sku, code, name, supplier, construction, game, old_id)
    # ─── Кий цельный / Пирамида ───
    ("111-01", "PR158",  "Porter Ramin 158 см",          "Мейлин", "Цельный",       "Пирамида", "5865"),
    ("111-16", "CM161",  "Compositor 161 см",            "Smart",  "Цельный",       "Пирамида", "9201"),
    ("111-17", "CM122",  "Compositor укороч. 122 см",    "Smart",  "Цельный",       "Пирамида", "15549"),
    ("111-18", "CM133",  "Compositor укороч. 133 см",    "Smart",  "Цельный",       "Пирамида", "15550"),
    ("111-19", "TG158",  "Тафгай карбон 158 см",        "Smart",  "Цельный",       "Пирамида", "16505"),
    ("111-26", "CR1",    "Crown 161 см (вар.1)",         "Конди",  "Цельный",       "Пирамида", "16276"),
    ("111-27", "CR2",    "Crown 161 см (вар.2)",         "Конди",  "Цельный",       "Пирамида", "13203"),
    # ─── Кий цельный / Пул ───
    ("112-01", "PR150",  "Porter Ramin 150 см",          "Мейлин", "Цельный",       "Пул",      "5866"),
    ("112-16", "CM148",  "Compositor 148 см",            "Smart",  "Цельный",       "Пул",      "15152"),
    ("112-17", "TG147",  "Тафгай карбон 147 см",        "Smart",  "Цельный",       "Пул",      "16506"),
    ("112-18", "TG122",  "Тафгай укороч. 122 см",       "Smart",  "Цельный",       "Пул",      "16508"),
    ("112-19", "TG133",  "Тафгай укороч. 133 см",       "Smart",  "Цельный",       "Пул",      "16507"),
    # ─── Кий цельный / Снукер ───
    ("113-01", "OXF",    "Oxford",                       "Мейлин", "Цельный",       "Снукер",   "5398"),
    ("113-02", "DND145", "Dandy 145 см",                 "Мейлин", "Цельный",       "Снукер",   "16518"),
    ("113-03", "BRJ143", "Bridge JDH 143 см",            "Мейлин", "Цельный",       "Снукер",   "16419"),
    # ─── Кий разборный / Пирамида ───
    ("121-26", "D501R",  "Maple Crown D-501R 163 см",    "Конди",  "Разборный (2ч)","Пирамида", "9334"),
    ("121-27", "D502R",  "Maple Crown D-502R 163 см",    "Конди",  "Разборный (2ч)","Пирамида", "9335"),
    ("121-28", "D510R",  "Maple Crown D-510R 163 см",    "Конди",  "Разборный (2ч)","Пирамида", "9336"),
    # ─── Кий разборный / Пул ───
    ("122-01", "PR146",  "Porter Ramin 146 см",          "Мейлин", "Разборный (2ч)","Пул",      "16230"),
    ("122-26", "D501",   "Maple Crown D-501 149 см",     "Конди",  "Разборный (2ч)","Пул",      "9339"),
    ("122-27", "D509",   "Maple Crown D-509 149 см",     "Конди",  "Разборный (2ч)","Пул",      "16519"),
    ("122-28", "D510",   "Maple Crown D-510 149 см",     "Конди",  "Разборный (2ч)","Пул",      "9340"),
    ("122-29", "MA01",   "Maple Crown MA-01 151 см",     "Конди",  "Разборный (2ч)","Пул",      "16501"),
    ("122-30", "MA02",   "Maple Crown MA-02 151 см",     "Конди",  "Разборный (2ч)","Пул",      "16502"),
    ("122-31", "MA03",   "Maple Crown MA-03 151 см",     "Конди",  "Разборный (2ч)","Пул",      "16503"),
    # ─── Кий разборный / Снукер ───
    ("123-01", "BR143",  "Bridge 3/4 143 см",            "Мейлин", "Разборный (2ч)","Снукер",   "16418"),
    ("123-02", "MAN",    "Manchester 3/4",                "Мейлин", "Разборный (2ч)","Снукер",   "9157"),
    ("123-03", "LIV1",   "Liverpool (вар.1)",             "Мейлин", "Разборный (2ч)","Снукер",   "12490"),
    ("123-04", "LIV2",   "Liverpool (вар.2)",             "Мейлин", "Разборный (2ч)","Снукер",   "9405"),
    ("123-05", "BRD",    "Bradford 3/4",                  "Мейлин", "Разборный (2ч)","Снукер",   "12487"),
    ("123-06", "BRDB",   "Bradford Black 3/4",            "Мейлин", "Разборный (2ч)","Снукер",   "13601"),
]

# Supplier colors
sup_colors = {
    "Мейлин": "E8F5E9",
    "Smart":  "E3F2FD",
    "Конди":  "FFF3E0",
}

prev_group = ""
row = 2
for sku, code, name, supplier, construction, game, old_id in cues:
    group = f"{construction} / {game}"
    if group != prev_group:
        # Insert group separator row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        cell = ws.cell(row=row, column=1, value=f"── {group} ──")
        cell.font = cat_font
        cell.fill = cat_fill
        cell.alignment = Alignment(horizontal='left')
        row += 1
        prev_group = group

    ws.cell(row=row, column=1, value=sku).font = sku_font
    c = ws.cell(row=row, column=2, value=code)
    c.font = code_font
    ws.cell(row=row, column=3, value=name)
    sup_cell = ws.cell(row=row, column=4, value=supplier)
    if supplier in sup_colors:
        sup_cell.fill = PatternFill("solid", fgColor=sup_colors[supplier])
    ws.cell(row=row, column=5, value=construction)
    ws.cell(row=row, column=6, value=game)
    ws.cell(row=row, column=7, value=old_id)

    for col in range(1, 8):
        ws.cell(row=row, column=col).border = thin_border
        ws.cell(row=row, column=col).alignment = Alignment(vertical='center')
    row += 1

# Freeze top row
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:G{row-1}"

out = "c:\\Projects\\ERP\\cues_codes.xlsx"
wb.save(out)
print(f"Saved: {out} ({row-2} items)")
