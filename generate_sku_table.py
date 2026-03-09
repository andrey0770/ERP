import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Категории артикулов"

# Styles
header_font = Font(bold=True, size=12)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font_white = Font(bold=True, size=11, color="FFFFFF")
cat_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
cat_font = Font(bold=True, size=11)
reserve_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
reserve_font = Font(italic=True, color="999999")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# Headers
headers = ["Код (CC)", "Категория товара", "Товаров сейчас", "Резерв (доп. позиции)", "Атрибуты (поз. 3-4)", "Примечание"]
ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 40
ws.column_dimensions['C'].width = 16
ws.column_dimensions['D'].width = 22
ws.column_dimensions['E'].width = 30
ws.column_dimensions['F'].width = 40

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

# Category data: (code, name, count, attributes, note)
categories = {
    # Кии (01-09)
    1: ("Кий для пирамиды", "~15", "[части][материал]", ""),
    2: ("Кий для пула", "~18", "[части][материал]", ""),
    3: ("Кий для снукера", "~9", "[части][материал]", ""),
    4: ("Кий универсальный", "0", "[части][материал]", "резерв"),
    5: ("Кий укороченный", "~6", "[части][материал]", "пирамида + пул"),

    # Шары (10-19)
    10: ("Шары для пирамиды", "0", "00", ""),
    11: ("Шары для пула", "~1", "00", ""),
    12: ("Шары для снукера", "~1", "00", ""),
    13: ("Шары тренировочные", "0", "00", ""),

    # Сукно (20-29)
    20: ("Сукно", "~60", "00", "самая большая группа в этом блоке"),

    # Чехлы и тубусы (30-39)
    30: ("Чехол для стола", "~69", "00", "покрывала и чехлы"),
    31: ("Тубус для кия", "~19", "00", "тубусы и футляры"),
    32: ("Чехол для кия", "0", "00", "мягкие чехлы"),

    # Аксессуары (40-49)
    40: ("Перчатки", "~19", "00", ""),
    41: ("Мел", "~3", "00", ""),
    42: ("Наклейки для киев", "~17", "00", "наклейки, инструмент"),
    43: ("Треугольники", "0", "00", ""),
    44: ("Щётки для сукна", "0", "00", ""),
    45: ("Полки / киевницы", "0", "00", ""),

    # Комплектующие столов (50-59)
    50: ("Лузы", "~16", "00", ""),
    51: ("Резина для бортов", "~11", "00", ""),
    52: ("Сетки для луз", "0", "00", ""),

    # Освещение (60-69)
    60: ("Светильники бильярдные", "0", "00", ""),

    # Столы (70-79)
    70: ("Стол бильярдный", "0", "00", ""),

    # Прочие аксессуары бильярд (80-89)
    80: ("Аксессуары прочие", "~170", "00", "основная «свалка», можно разбить"),

    # Другие виды спорта (90-99)
    90: ("Карты игральные", "~1", "00", "покер и др."),
    91: ("Мишени и дротики", "0", "00", "дартс"),
    92: ("Ракетки и мячи н/т", "0", "00", "настольный теннис"),
    93: ("Игровой стол многофункц.", "0", "00", ""),
    94: ("Тренажёры", "0", "00", ""),
}

# Fill all 100 rows (00-99)
for code in range(100):
    row = code + 2  # row 2 = code 00, row 101 = code 99
    code_str = f"{code:02d}"

    ws.cell(row=row, column=1, value=code_str).border = thin_border
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')

    if code in categories:
        name, count, attrs, note = categories[code]
        ws.cell(row=row, column=2, value=name).border = thin_border
        ws.cell(row=row, column=3, value=count).border = thin_border
        ws.cell(row=row, column=4, value="").border = thin_border  # user fills this
        ws.cell(row=row, column=5, value=attrs).border = thin_border
        ws.cell(row=row, column=6, value=note).border = thin_border

        # Style category rows
        for col in range(1, 7):
            ws.cell(row=row, column=col).fill = cat_fill
            ws.cell(row=row, column=col).font = cat_font
    else:
        ws.cell(row=row, column=2, value="— резерв —").border = thin_border
        ws.cell(row=row, column=3, value="").border = thin_border
        ws.cell(row=row, column=4, value="").border = thin_border
        ws.cell(row=row, column=5, value="").border = thin_border
        ws.cell(row=row, column=6, value="").border = thin_border

        # Style reserve rows
        for col in range(1, 7):
            ws.cell(row=row, column=col).fill = reserve_fill
            ws.cell(row=row, column=col).font = reserve_font

    ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')
    ws.cell(row=row, column=4).alignment = Alignment(horizontal='center')

# Freeze header
ws.freeze_panes = 'A2'

# Second sheet: legend for cue attributes
ws2 = wb.create_sheet("Атрибуты киев")
ws2.column_dimensions['A'].width = 15
ws2.column_dimensions['B'].width = 30
ws2.column_dimensions['C'].width = 15
ws2.column_dimensions['D'].width = 30

for col, h in enumerate(["Поз. 3", "Кол-во частей", "Поз. 4", "Материал"], 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.border = thin_border

parts = [(1, "Цельный"), (2, "Разборный (2 части)"), (3, "3/4 (3 части)")]
materials = [(1, "Рамин"), (2, "Клён"), (3, "Композит / карбон")]

for i, (code, name) in enumerate(parts, 2):
    ws2.cell(row=i, column=1, value=code).border = thin_border
    ws2.cell(row=i, column=2, value=name).border = thin_border

for i, (code, name) in enumerate(materials, 2):
    ws2.cell(row=i, column=3, value=code).border = thin_border
    ws2.cell(row=i, column=4, value=name).border = thin_border

# Third sheet: example
ws3 = wb.create_sheet("Пример — Кии")
ws3.column_dimensions['A'].width = 12
ws3.column_dimensions['B'].width = 55
ws3.column_dimensions['C'].width = 15
ws3.column_dimensions['D'].width = 45

for col, h in enumerate(["Артикул", "Товар", "Старый SKU", "Расшифровка"], 1):
    cell = ws3.cell(row=1, column=col, value=h)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.border = thin_border

examples = [
    ("011101", "Porter Ramin цельный 158 см", "5865", "пирамида | цельный | рамин | №01"),
    ("011301", "Compositor цельный 161 см", "9201", "пирамида | цельный | композит | №01"),
    ("011201", "Crown цельный 161 см", "16276", "пирамида | цельный | клён | №01"),
    ("011202", "Player Black цельный 159 см", "13202", "пирамида | цельный | клён | №02"),
    ("011203", "Player Brown цельный 159 см", "15874", "пирамида | цельный | клён | №03"),
    ("011301", "Тафгай карбоновый цельный 158 см", "16505", "пирамида | цельный | композит | №01"),
    ("012201", "Maple Crown D-501R разборный 163 см", "9334", "пирамида | разборный | клён | №01"),
    ("012202", "Maple Crown D-502R разборный 163 см", "9335", "пирамида | разборный | клён | №02"),
    ("012203", "Maple Crown D-510R разборный 163 см", "9336", "пирамида | разборный | клён | №03"),
    ("012204", "Astro Atlas разборный 161 см", "12466", "пирамида | разборный | клён | №04"),
    ("012205", "Astro Excalibur разборный 161 см", "9270", "пирамида | разборный | клён | №05"),
    ("012206", "Astro Hercules разборный 161 см", "9269", "пирамида | разборный | клён | №06"),
    ("", "", "", ""),
    ("021101", "Porter Ramin цельный 150 см", "5866", "пул | цельный | рамин | №01"),
    ("021301", "Compositor цельный 148 см", "15152", "пул | цельный | композит | №01"),
    ("022201", "Maple Crown D-501 разборный 149 см", "9339", "пул | разборный | клён | №01"),
    ("022202", "Maple Crown D-509 разборный 149 см", "16519", "пул | разборный | клён | №02"),
    ("022203", "Maple Crown D-510 разборный 149 см", "9340", "пул | разборный | клён | №03"),
    ("022204", "Astro Atlas разборный 148 см", "9139", "пул | разборный | клён | №04"),
    ("", "", "", ""),
    ("031201", "Oxford цельный", "5398", "снукер | цельный | клён | №01"),
    ("032201", "Manchester разборный 3/4", "9157", "снукер | разборный | клён | №01"),
    ("033201", "Bradford разборный 3/4", "12487", "снукер | 3/4 | клён | №01"),
    ("", "", "", ""),
    ("051301", "Compositor укороченный 122 см", "15549", "укороченный | цельный | композит | №01"),
    ("051302", "Compositor укороченный 133 см", "15550", "укороченный | цельный | композит | №02"),
    ("", "", "", ""),
    ("410001", "Мел Master синий 4шт", "9536-4", "мел | 00 | №01"),
    ("200001", "Сукно Iwan Simonis 760", "", "сукно | 00 | №01"),
]

for i, (sku, name, old, desc) in enumerate(examples, 2):
    ws3.cell(row=i, column=1, value=sku).border = thin_border
    ws3.cell(row=i, column=2, value=name).border = thin_border
    ws3.cell(row=i, column=3, value=old).border = thin_border
    ws3.cell(row=i, column=4, value=desc).border = thin_border
    ws3.cell(row=i, column=1).font = Font(bold=True, name="Consolas", size=11)

out = "sku_categories.xlsx"
wb.save(out)
print(f"Saved: {out}")
